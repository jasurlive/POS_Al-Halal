import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from datetime import datetime


class ExcelHandler:
    def __init__(self):
        self.file_path = self.get_excel_path()
        self.create_excel_if_not_exists()

    def get_excel_path(self):
        """Generate file path based on the current year and month."""
        current_year = datetime.now().year
        current_month = datetime.now().month
        file_name = f"POS_{current_year}_{current_month:02d}.xlsx"
        return os.path.join("data", file_name)

    def create_excel_if_not_exists(self):
        """Create a new Excel file with the necessary structure if it does not exist."""
        if not os.path.exists(self.file_path):
            wb = Workbook()
            month_name = datetime.now().strftime("%B")
            ws = wb.active
            ws.title = month_name

            # Define headers
            headers = [
                "No",
                "Barcode",
                "Item Name",
                "Inventory Quantity",
                "Quantity Sold",
                "Quantity Left",
                "Original Price",
                "Sale Price",
                "Total Profit",
                "Invested",
                "Clean Profit",
                None,
            ]

            ws.append(headers)

            # Apply alignment and style to headers
            for cell in ws[1]:
                self.apply_formatting(cell, "header")

            # Add formula templates for first 100 rows
            for row in range(2, 102):
                ws[f"I{row}"] = f"=H{row}*E{row}"
                ws[f"J{row}"] = f"=D{row}*G{row}"
                ws[f"K{row}"] = f"=I{row}-J{row}"

            wb.save(self.file_path)

    def load_workbook(self):
        """Load the workbook."""
        return load_workbook(self.file_path)

    def save_workbook(self, wb):
        """Save the workbook."""
        wb.save(self.file_path)

    def get_column_indexes(self, ws):
        """Retrieve column indexes based on header names."""
        try:
            headers = {
                ws.cell(row=1, column=col).value: col
                for col in range(1, ws.max_column + 1)
            }
            required_columns = [
                "Barcode",
                "Item Name",
                "Inventory Quantity",
                "Original Price",
                "Sale Price",
            ]

            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                raise ValueError(
                    f"Missing required columns in Excel: {missing_columns}"
                )

            return {col: headers[col] for col in required_columns}

        except Exception as e:
            print(f"Error getting column indexes: {e}")
            return {}

    def find_first_empty_row(self, ws, barcode_col):
        """Find the first available empty row in the Barcode column."""
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip headers)
            if ws.cell(row=row, column=barcode_col).value is None:
                return row
        return ws.max_row + 1  # If no empty row found, append at the end

    def apply_formatting(self, cell, cell_type):
        """Apply alignment and formatting to a cell."""
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell_type == "header":
            cell.font = Font(name="Calibri", size=11, bold=True)
        else:
            cell.font = Font(name="Calibri", size=11)


# Instantiate the ExcelHandler to create or load the Excel file
excel_handler = ExcelHandler()
