import pandas as pd
import os
from datetime import datetime
import openpyxl


class POSHandler:
    def __init__(self):
        self.file_path = self.get_excel_path()

    def get_excel_path(self):
        current_year = datetime.now().year
        current_month = datetime.now().month
        file_name = f"POS_{current_year}_{current_month:02d}.xlsx"
        return os.path.join("data", file_name)

    def find_item_by_barcode(self, barcode):
        df = pd.read_excel(self.file_path, dtype={"Barcode": str}, engine="openpyxl")
        item = df[df["Barcode"] == barcode]
        return item.iloc[0].to_dict() if not item.empty else None

    def update_inventory(self, item_name):
        """Update the inventory by reducing the quantity of the sold item."""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        wb = openpyxl.load_workbook(self.file_path)
        ws = wb.active

        # Get column indexes dynamically
        columns = self.get_column_indexes(ws)

        # Search for the item in the "Item Name" column
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip headers)
            if ws.cell(row=row, column=columns["Item Name"]).value == item_name:
                inventory_cell = ws.cell(row=row, column=columns["Inventory Quantity"])
                # Ensure inventory value is numeric
                try:
                    current_quantity = int(inventory_cell.value)
                except (ValueError, TypeError):
                    current_quantity = 0
                inventory_cell.value = max(0, current_quantity - 1)
                break

        # Save workbook
        wb.save(self.file_path)

    def get_column_indexes(self, ws):
        """Find column indexes based on header names."""
        headers = {
            ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)
        }
        required_columns = [
            "Barcode",
            "Item Name",
            "Inventory Quantity",
            "Original Price",
            "Sale Price",
        ]

        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            raise ValueError(f"Missing required columns in Excel: {missing_columns}")

        return {col: headers[col] for col in required_columns}
