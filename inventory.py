import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


class InventoryHandler:
    def __init__(self):
        self.file_path = self.get_excel_path()

    def get_excel_path(self):
        current_year = datetime.now().year
        current_month = datetime.now().month
        file_name = f"POS_{current_year}_{current_month:02d}.xlsx"
        return os.path.join("data", file_name)

    def get_column_indexes(self, ws):
        """Find column indexes based on header names."""
        headers = {
            ws.cell(row=1, column=col).value: col for col in range(1, ws.max_column + 1)
        }
        required_columns = [
            "Barcode",
            "Item Name",
            "Inventory Quantity",
            "Original Price",
            "Sale Price",
        ]

        missing_columns = [col for col in required_columns if col not in headers]
        if missing_columns:
            raise ValueError(f"Missing required columns in Excel: {missing_columns}")

        return {col: headers[col] for col in required_columns}

    def find_first_empty_row(self, ws, barcode_col):
        """Find the first available empty row in the Barcode column."""
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip headers)
            if ws.cell(row=row, column=barcode_col).value is None:
                return row
        return ws.max_row + 1  # If no empty row found, append at the end

    def add_inventory_item(
        self,
        barcode,
        item_name,
        original_price,
        sale_price,
        inventory_quantity,
    ):
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        # Validate and sanitize inputs
        barcode = str(barcode).strip()
        item_name = str(item_name).strip()
        try:
            original_price = float(original_price)
        except ValueError:
            raise ValueError("Original Price must be a valid number.")
        try:
            sale_price = float(sale_price)
        except ValueError:
            raise ValueError("Sale Price must be a valid number.")
        try:
            inventory_quantity = int(inventory_quantity)
        except ValueError:
            raise ValueError("Inventory Quantity must be a valid integer.")

        wb = load_workbook(self.file_path)
        ws = wb.active

        # Get column indexes dynamically
        columns = self.get_column_indexes(ws)

        # Check if the barcode already exists
        existing_row = None
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip headers)
            if str(ws.cell(row=row, column=columns["Barcode"]).value) == barcode:
                existing_row = row
                break

        # If barcode exists, update the row; otherwise, find the first empty row
        target_row = (
            existing_row
            if existing_row
            else self.find_first_empty_row(ws, columns["Barcode"])
        )

        # Prepare data to insert
        data = {
            "Barcode": barcode,
            "Item Name": item_name,
            "Inventory Quantity": inventory_quantity,
            "Original Price": original_price,
            "Sale Price": sale_price,
        }

        # Insert or update data into the correct columns dynamically
        for column_name, value in data.items():
            cell = ws.cell(row=target_row, column=columns[column_name], value=value)
            # Apply alignment and formatting
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if column_name in ["Original Price", "Sale Price"]:
                cell.number_format = "#,##0.00"  # Format as currency
            cell.font = Font(name="Calibri", size=11)

        # Save workbook
        wb.save(self.file_path)

    def get_inventory_item(self, barcode):
        """Fetch an item's details by barcode."""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        wb = load_workbook(self.file_path)
        ws = wb.active

        # Get column indexes dynamically
        columns = self.get_column_indexes(ws)

        # Search for the barcode in the "Barcode" column
        for row in range(2, ws.max_row + 1):  # Start from row 2 (skip headers)
            if str(ws.cell(row=row, column=columns["Barcode"]).value) == str(barcode):
                return {
                    column_name: ws.cell(row=row, column=col_index).value
                    for column_name, col_index in columns.items()
                }

        return None  # Return None if the item is not found
